#!/usr/bin/env python3
"""Merge store PLU spreadsheets into Larder plu-reference.json (+ CSV).

Sources (data/work-from-store/):
  plu/*.xlsx department lists, code books
  costs/Copy of PLU Costs.xlsx

Also seeds from existing server/src/static/plu-reference.json.

Usage:
  python3 scripts/merge_plu_reference.py
"""

from __future__ import annotations

import csv
import json
import re
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from xml.etree import ElementTree

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "data" / "work-from-store"
EXISTING = ROOT / "server" / "src" / "static" / "plu-reference.json"
OUT_JSON = EXISTING
OUT_CSV = ROOT / "server" / "src" / "static" / "plu-reference.csv"
OUT_CLEAN_JSON = ROOT / "data" / "lexco-source" / "clean" / "plu-reference.json"
OUT_CLEAN_CSV = ROOT / "data" / "lexco-source" / "clean" / "plu-reference.csv"
OUT_REPORT = ROOT / "data" / "work-from-store" / "plu-merge-report.txt"

# Prepared-foods depts beat generic "existing" so re-runs refresh dept/price.
DEPT_PRIORITY = {
    "deli": 20,
    "bakehouse": 20,
    "bakery": 20,
    "prepared": 18,
    "code-book": 16,
    "cost-sheet": 14,
    "shrink": 12,
    "hot-bar": 12,
    "turkey": 10,
    "cheese": 8,
    "meat": 7,
    "fish": 7,
    "bulk": 6,
    "small-vendor": 6,
    "local-produce": 6,
    "produce": 6,
    "existing": 4,  # seed only; store sheets win on conflict
    "other": 1,
}


def clean_name(s: object) -> str:
    if s is None:
        return ""
    t = str(s).replace("\xa0", " ").strip()
    t = re.sub(r"\s+", " ", t)
    return t


def normalize_plu(val: object) -> str | None:
    if val is None:
        return None
    if isinstance(val, float):
        if val != val:  # NaN
            return None
        if val == int(val):
            val = int(val)
    s = str(val).strip()
    if not s or s.lower() in {"need plu", "new", "n/a", "none", "#n/a", "see below"}:
        return None
    # strip .0 from floats-as-str
    if re.fullmatch(r"\d+\.0+", s):
        s = s.split(".")[0]
    s = re.sub(r"[^\d]", "", s)
    if not s:
        return None
    # Skip full UPC/EAN barcodes (prepared PLUs are short)
    if len(s) >= 11:
        return None
    if len(s) < 3 or len(s) > 6:
        return None
    return s


def add_item(
    by_key: dict[str, dict],
    *,
    name: str,
    plu: str,
    source: str,
    dept: str = "",
    price: object = None,
    size: str = "",
) -> None:
    name = clean_name(name)
    plu = normalize_plu(plu)
    if not name or not plu:
        return
    # skip pure section headers
    if name.endswith(":") and len(name) < 40:
        return

    # A product name can legitimately carry multiple PLUs (sizes, seasonal
    # versions, or department-specific codes).  Keep each name/code pair;
    # only competing details for the exact same pair are reconciled.
    key = (name.casefold(), plu)
    pri = DEPT_PRIORITY.get(dept.casefold(), 1)
    price_s = ""
    if price is not None and price != "" and not isinstance(price, datetime_type()):
        try:
            if isinstance(price, (int, float)) and price == price:
                price_s = f"{float(price):.2f}".rstrip("0").rstrip(".")
            else:
                ps = str(price).strip()
                if re.match(r"^[\d.]+$", ps):
                    price_s = ps
        except Exception:
            price_s = ""

    item = {
        "name": name,
        "plu": plu,
        "saleable": "0",
        "author": "",
        "source": source,
        "dept": clean_name(dept).casefold() if dept else "",
        "size": clean_name(size),
        "price": price_s,
        "_pri": pri,
    }
    prev = by_key.get(key)
    if prev is None:
        by_key[key] = item
        return
    # Prefer higher dept priority; if same, keep existing plu unless empty
    if item["_pri"] > prev.get("_pri", 0):
        # keep name casing from higher priority source
        by_key[key] = item
    elif item["_pri"] == prev.get("_pri", 0) and prev.get("plu") != plu:
        # conflict note — keep first, track alt
        alts = prev.setdefault("_alt_plus", set())
        if isinstance(alts, set):
            alts.add(plu)
        if not prev.get("price") and price_s:
            prev["price"] = price_s
    else:
        if not prev.get("price") and price_s:
            prev["price"] = price_s
        if not prev.get("dept") and item.get("dept"):
            prev["dept"] = item["dept"]


def datetime_type():
    from datetime import datetime

    return datetime


def iter_sheet_rows(path: Path, sheet: str | None = None):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        names = [sheet] if sheet else wb.sheetnames
        for sn in names:
            if sn not in wb.sheetnames:
                continue
            ws = wb[sn]
            for row in ws.iter_rows(values_only=True):
                yield sn, row
    finally:
        wb.close()


def load_existing(by_key: dict) -> int:
    if not EXISTING.is_file():
        return 0
    data = json.loads(EXISTING.read_text(encoding="utf-8"))
    n = 0
    for it in data.get("items") or []:
        add_item(
            by_key,
            name=it.get("name") or "",
            plu=it.get("plu") or "",
            source=it.get("source") or "existing",
            dept="existing",
        )
        n += 1
    return n


def parse_code_book(path: Path, by_key: dict) -> int:
    n = 0
    for _sn, row in iter_sheet_rows(path):
        if not row:
            continue
        cells = [c for c in row if c is not None]
        if len(cells) < 2:
            continue
        # name, plu  OR  section header
        name, plu = cells[0], cells[1]
        if normalize_plu(plu):
            add_item(by_key, name=str(name), plu=plu, source=path.name, dept="code-book")
            n += 1
    return n


def parse_headered(
    path: Path,
    by_key: dict,
    *,
    sheet: str | None,
    dept: str,
    name_keys: list[str],
    plu_keys: list[str],
    price_keys: list[str] | None = None,
    size_keys: list[str] | None = None,
    dept_keys: list[str] | None = None,
) -> int:
    """Find header row containing plu+name, then parse."""
    price_keys = price_keys or []
    size_keys = size_keys or []
    dept_keys = dept_keys or []
    n = 0
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = [sheet] if sheet else wb.sheetnames
        for sn in sheets:
            if sn not in wb.sheetnames:
                continue
            ws = wb[sn]
            header_idx = None
            headers: list[str] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                cells = list(row or ())
                labels = [clean_name(c).casefold() for c in cells]
                if any(k.casefold() in labels for k in plu_keys) and any(
                    k.casefold() in labels for k in name_keys
                ):
                    header_idx = i
                    headers = labels
                    break
            if header_idx is None:
                # fallback: first row with PLU + Description text
                continue

            def col(keys: list[str]) -> int | None:
                for k in keys:
                    kk = k.casefold()
                    if kk in headers:
                        return headers.index(kk)
                return None

            i_plu = col(plu_keys)
            i_name = col(name_keys)
            i_price = col(price_keys) if price_keys else None
            i_size = col(size_keys) if size_keys else None
            i_dept = col(dept_keys) if dept_keys else None
            if i_plu is None or i_name is None:
                continue

            for j, row in enumerate(ws.iter_rows(values_only=True)):
                if j <= header_idx:
                    continue
                cells = list(row or ())
                def get(i):
                    return cells[i] if i is not None and i < len(cells) else None

                plu = get(i_plu)
                name = get(i_name)
                # meat sheet sometimes has brand as col2 and desc as col3; name col may be empty
                if not clean_name(name) and len(cells) > 2:
                    # try combine brand+description if description is longer
                    pass
                row_dept = clean_name(get(i_dept)) if i_dept is not None else dept
                if not row_dept:
                    row_dept = dept
                add_item(
                    by_key,
                    name=str(name or ""),
                    plu=plu,
                    source=f"{path.name}:{sn}",
                    dept=row_dept or dept,
                    price=get(i_price),
                    size=str(get(i_size) or ""),
                )
                if normalize_plu(plu) and clean_name(name):
                    n += 1
    finally:
        wb.close()
    return n


def parse_deli_bakehouse_loose(path: Path, by_key: dict, dept: str) -> int:
    """Deli master: sometimes header PLU/Type/Description/Price with messy columns."""
    n = 0
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for sn in wb.sheetnames:
            ws = wb[sn]
            # skip POS dump sheets
            if sn.lower() in {"sheet1", "sheet2", "db", "test"} and "list" not in path.name.casefold():
                # still try print list / master
                pass
            for row in ws.iter_rows(values_only=True):
                cells = list(row or ())
                if not cells:
                    continue
                # find first numeric short PLU in row
                plu_i = None
                plu_v = None
                for i, c in enumerate(cells):
                    p = normalize_plu(c)
                    if p:
                        plu_i = i
                        plu_v = p
                        break
                if plu_i is None:
                    continue
                # name: prefer next non-empty string that isn't a price
                name = ""
                price = None
                for i, c in enumerate(cells):
                    if i == plu_i:
                        continue
                    if c is None:
                        continue
                    if isinstance(c, (int, float)) and not isinstance(c, bool):
                        if c == c and 0 < float(c) < 500:
                            price = c
                        continue
                    s = clean_name(c)
                    if not s or s.casefold() in {"plu", "type", "description", "price", "category", "sub category", "retail"}:
                        continue
                    if re.fullmatch(r"[\d.]+", s):
                        continue
                    # skip date-like
                    if "datetime" in type(c).__name__.lower():
                        continue
                    if len(s) > len(name):
                        name = s
                if name and plu_v:
                    # filter out personal care / grocery noise from dump sheets
                    low = name.casefold()
                    if any(x in low for x in ("shampoo", "lotion", "conditioner", "sanitizer", "bubble bath")):
                        continue
                    add_item(
                        by_key,
                        name=name,
                        plu=plu_v,
                        source=f"{path.name}:{sn}",
                        dept=dept,
                        price=price,
                    )
                    n += 1
    finally:
        wb.close()
    return n


def parse_meat_fish(path: Path, by_key: dict) -> int:
    n = 0
    for sn, row in iter_sheet_rows(path):
        if sn.casefold() not in {"meat", "fish"}:
            continue
        cells = list(row or ())
        if not cells:
            continue
        plu = normalize_plu(cells[0])
        if not plu:
            continue
        # brand, description, size — description may be col1 or col2
        parts = [clean_name(c) for c in cells[1:] if clean_name(c) and "datetime" not in type(c).__name__.lower()]
        # drop size-only trailing LB
        name_parts = [p for p in parts if p.casefold() not in {"lb", "ct", "1 ct", "12 ct"}]
        if not name_parts:
            continue
        # if first is brand-ish and second is product, join
        name = " ".join(name_parts[:2]) if len(name_parts) >= 2 else name_parts[0]
        add_item(by_key, name=name, plu=plu, source=f"{path.name}:{sn}", dept=sn.casefold())
        n += 1
    return n


def parse_shrink(path: Path, by_key: dict) -> int:
    n = 0
    for sn, row in iter_sheet_rows(path):
        cells = [c for c in (row or ()) if c is not None]
        if len(cells) < 2:
            continue
        # Item, PLU
        name, plu = cells[0], cells[1]
        if normalize_plu(plu) and clean_name(name) and clean_name(name).casefold() not in {"item", "plu"}:
            dept = "shrink"
            if "bake" in sn.casefold():
                dept = "bakehouse"
            elif "deli" in sn.casefold() or "pfd" in sn.casefold():
                dept = "deli"
            add_item(by_key, name=str(name), plu=plu, source=f"{path.name}:{sn}", dept=dept)
            n += 1
    return n


def parse_turkey(path: Path, by_key: dict) -> int:
    n = 0
    for sn, row in iter_sheet_rows(path):
        cells = list(row or ())
        if len(cells) < 2:
            continue
        plu = normalize_plu(cells[0])
        name = clean_name(cells[1])
        if plu and name and name.casefold() not in {"brand", "item id"}:
            price = cells[3] if len(cells) > 3 else None
            add_item(by_key, name=name, plu=plu, source=path.name, dept="turkey", price=price)
            n += 1
    return n


def parse_small_vendor(path: Path, by_key: dict) -> int:
    """Parse the three side-by-side name/PLU columns in the vendor guide."""
    n = 0
    for _sn, row in iter_sheet_rows(path):
        cells = list(row or ())
        for i, cell in enumerate(cells):
            plu = normalize_plu(cell)
            if not plu or i == 0:
                continue
            name = clean_name(cells[i - 1])
            if not name:
                continue
            add_item(
                by_key,
                name=name,
                plu=plu,
                source=path.name,
                dept="small-vendor",
            )
            n += 1
    return n


def parse_local_produce(path: Path, by_key: dict) -> int:
    """Read the DOCX's text and collect `item – PLU` lines."""
    try:
        with zipfile.ZipFile(path) as docx:
            root = ElementTree.fromstring(docx.read("word/document.xml"))
    except (KeyError, OSError, zipfile.BadZipFile, ElementTree.ParseError):
        return 0
    text = "\n".join(t.text or "" for t in root.iter("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t"))
    n = 0
    for name, plu in re.findall(r"([^\n–—]+?)\s*[–—]\s*(\d{3,6})(?=\s|$)", text):
        add_item(
            by_key,
            name=name,
            plu=plu,
            source=path.name,
            dept="local-produce",
        )
        n += 1
    return n


def parse_produce_inventory(path: Path, by_key: dict) -> int:
    """Parse the POS export: blank column, alternate ID, name, price, note.

    Full UPC/EAN values are deliberately skipped by ``normalize_plu``; short
    alternate IDs are store PLUs and belong in the lookup.
    """
    n = 0
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as f:
        for row in csv.reader(f):
            if len(row) < 3:
                continue
            plu = normalize_plu(row[1])
            name = clean_name(row[2])
            if not plu or not name:
                continue
            price = clean_name(row[3]).lstrip("$") if len(row) > 3 else ""
            add_item(
                by_key,
                name=name,
                plu=plu,
                source=path.name,
                dept="produce",
                price=price,
            )
            n += 1
    return n


def main() -> None:
    by_key: dict[str, dict] = {}
    report: list[str] = []

    n = load_existing(by_key)
    report.append(f"existing reference: {n} rows → {len(by_key)} unique names")

    plu_dir = SRC / "plu"
    costs = SRC / "costs" / "Copy of PLU Costs.xlsx"

    jobs = [
        ("Code Book PLUs.xlsx", lambda p: parse_code_book(p, by_key)),
        (
            "PLU List Deli.xlsx",
            lambda p: parse_headered(
                p,
                by_key,
                sheet="master",
                dept="deli",
                name_keys=["description", "name"],
                plu_keys=["plu", "scan code", "scancode"],
                price_keys=["price", "retail", "baseprice", "base price"],
            )
            or parse_deli_bakehouse_loose(p, by_key, "deli"),
        ),
        (
            "PLU List Bakehouse.xlsx",
            lambda p: parse_headered(
                p,
                by_key,
                sheet="print list",
                dept="bakehouse",
                name_keys=["description", "name"],
                plu_keys=["plu"],
                price_keys=["retail", "price"],
            )
            or parse_deli_bakehouse_loose(p, by_key, "bakehouse"),
        ),
        (
            "PLU LIST - Cheese.xlsx",
            lambda p: parse_headered(
                p,
                by_key,
                sheet="Master",
                dept="cheese",
                name_keys=["name"],
                plu_keys=["upc/plu", "plu", "upc"],
                price_keys=["price"],
                size_keys=["size"],
            ),
        ),
        ("PLU List Meat Fish.xlsx", lambda p: parse_meat_fish(p, by_key)),
        (
            "PLU BULK Repack Sheet Master.xlsx",
            lambda p: parse_headered(
                p,
                by_key,
                sheet="Master",
                dept="bulk",
                name_keys=["item", "name", "description"],
                plu_keys=["plu"],
                price_keys=["price"],
            ),
        ),
        ("Shrink Log Reference Bakehouse-Deli.xlsx", lambda p: parse_shrink(p, by_key)),
        ("Turkey PLU's.xlsx", lambda p: parse_turkey(p, by_key)),
        ("PLU List Small Vendor.xlsx", lambda p: parse_small_vendor(p, by_key)),
        ("Local Produce Codes, August 2017 JC.docx", lambda p: parse_local_produce(p, by_key)),
        ("produce-inventory-alt-id.csv", lambda p: parse_produce_inventory(p, by_key)),
    ]

    for fname, fn in jobs:
        path = plu_dir / fname
        if not path.is_file():
            report.append(f"skip missing {fname}")
            continue
        before = len(by_key)
        try:
            added = fn(path)
        except Exception as e:
            report.append(f"ERROR {fname}: {e}")
            continue
        report.append(f"{fname}: parsed~{added}, unique now {len(by_key)} (+{len(by_key)-before})")

    if costs.is_file():
        before = len(by_key)
        added = parse_headered(
            costs,
            by_key,
            sheet=None,
            dept="cost-sheet",
            name_keys=["name"],
            plu_keys=["scancode", "scan code", "plu"],
            price_keys=["baseprice", "base price", "lastcost", "last cost"],
            size_keys=["size"],
            dept_keys=["dep", "dept", "department"],
        )
        # costs often only Deli/Bakehouse short codes — also loose parse
        if added < 50:
            added = parse_deli_bakehouse_loose(costs, by_key, "cost-sheet")
        report.append(f"{costs.name}: parsed~{added}, unique now {len(by_key)} (+{len(by_key)-before})")

    # Common store typos / aliases → extra name keys pointing at same PLU
    ALIASES = {
        "anitpasta asparagus": "Asparagus Antipasto",
        "antipasto asparagus": "Asparagus Antipasto",
        "hummus classic": "Black Bean Hummus",  # often the house classic; also keep original
        "garlic hummus (classic)": "Garlic Hummus",
        "cookie cowgirl 4 ct": "Cowgirl Cookie",
        "cowgirl cookies": "Cowgirl Cookie",
    }
    # Build alias entries from existing names without overwriting stronger names.
    def find_by_name(name: str) -> dict | None:
        folded = name.casefold()
        for (item_name, _plu), item in by_key.items():
            if item_name == folded:
                return item
        return None

    extras: list[dict] = []
    for alias_key, display in ALIASES.items():
        # find source item by alias name or by display
        src = find_by_name(alias_key) or find_by_name(display)
        if not src:
            continue
        # ensure display name maps to this plu
        if not find_by_name(display):
            extras.append(
                {
                    "name": display,
                    "plu": src["plu"],
                    "saleable": "0",
                    "author": "",
                    "source": f"alias:{src.get('source','')}",
                    "dept": src.get("dept") or "",
                }
            )
        # ensure alias key maps if different spelling
        if not find_by_name(alias_key):
            extras.append(
                {
                    "name": display if alias_key.startswith(display.casefold()[:5]) else src["name"],
                    "plu": src["plu"],
                    "saleable": "0",
                    "author": "",
                    "source": f"alias:{src.get('source','')}",
                    "dept": src.get("dept") or "",
                }
            )

    items = []
    for it in list(by_key.values()) + extras:
        it = dict(it)
        it.pop("_pri", None)
        alts = it.pop("_alt_plus", None)
        if alts:
            it["alt_plus"] = sorted(alts)
        out = {
            "name": it["name"],
            "plu": it["plu"],
            "saleable": it.get("saleable") or "0",
            "author": it.get("author") or "",
            "source": it.get("source") or "",
        }
        if it.get("dept"):
            out["dept"] = it["dept"]
        if it.get("size"):
            out["size"] = it["size"]
        if it.get("price"):
            out["price"] = it["price"]
        if it.get("alt_plus"):
            out["alt_plus"] = it["alt_plus"]
        items.append(out)

    # de-dupe exact name+plu
    seen: set[tuple[str, str]] = set()
    deduped = []
    for it in items:
        k = (it["name"].casefold(), it["plu"])
        if k in seen:
            continue
        seen.add(k)
        deduped.append(it)
    items = deduped
    items.sort(key=lambda x: (x["name"].casefold(), x["plu"]))

    payload = {
        "title": "Lexington Co-op PLU reference",
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "count": len(items),
        "note": (
            "Merged from existing Larder reference + store PLU lists "
            "(Deli, Bakehouse, Cheese, Meat/Fish, Bulk, Small Vendor, Local Produce, "
            "Produce Inventory, Code Book, Costs, Shrink). "
            "Short store PLUs only (3–6 digits); full UPC barcodes omitted."
        ),
        "items": items,
    }

    for path in (OUT_JSON, OUT_CLEAN_JSON):
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
        report.append(f"wrote {path} ({len(items)} items)")

    for path in (OUT_CSV, OUT_CLEAN_CSV):
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(
                f,
                fieldnames=["name", "plu", "dept", "price", "size", "source"],
                extrasaction="ignore",
            )
            w.writeheader()
            for it in items:
                w.writerow(it)
        report.append(f"wrote {path}")

    # dept breakdown
    from collections import Counter

    c = Counter((it.get("dept") or "unknown") for it in items)
    report.append("by dept: " + ", ".join(f"{k}={v}" for k, v in c.most_common()))

    OUT_REPORT.write_text("\n".join(report) + "\n", encoding="utf-8")
    print("\n".join(report))
    print(f"\nTOTAL {len(items)} PLUs")


if __name__ == "__main__":
    main()
