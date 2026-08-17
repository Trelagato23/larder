#!/usr/bin/env python3
"""Clean LexCo / work-profile data into kitchen-friendly export packs.

Reads either:
  - a full larder bundle JSON (preferred), or
  - falls back to regenerating from data/lexco-source/lexco-bundle.json

Writes under data/lexco-source/clean/:
  kitchen-bundle.json     normalized full bundle (re-importable)
  recipes.csv             flat recipe lines (name, qty, unit, ingredient, …)
  ingredients-master.csv  unique ingredients + default unit + usage count
  by-dept/<dept>.md       markdown cookbooks per department
  by-dept/<dept>.json     recipe arrays per department
  index.md                TOC with counts
  sandwich-board.csv      from daily prep xlsx if present
  catalog.csv             RecpList enriched with imported yes/no

Usage:
  python3 scripts/clean_lexco_exports.py
  python3 scripts/clean_lexco_exports.py --bundle data/lexco-source/lexco-bundle.json
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "data" / "lexco-source"
OUT = SRC / "clean"

UNIT_MAP = {
    "tbl": "tbsp",
    "tbs": "tbsp",
    "tablespoon": "tbsp",
    "tablespoons": "tbsp",
    "tsp": "tsp",
    "teaspoon": "tsp",
    "teaspoons": "tsp",
    "cups": "cup",
    "cup": "cup",
    "c": "cup",
    "lbs": "lb",
    "pound": "lb",
    "pounds": "lb",
    "oz": "oz",
    "ounce": "oz",
    "ounces": "oz",
    "fl oz": "fl oz",
    "floz": "fl oz",
    "gal": "gal",
    "gallon": "gal",
    "gallons": "gal",
    "qt": "qt",
    "quart": "qt",
    "quarts": "qt",
    "pt": "pt",
    "pint": "pt",
    "ea": "ea",
    "each": "ea",
    "no unit": None,
    "none": None,
    "-": None,
    "": None,
}

# Collapse ChefTec category noise into kitchen-useful department tags
DEPT_ALIASES = {
    "bakery": "bakery",
    "soups": "soups",
    "soup": "soups",
    "sandwiches": "sandwiches",
    "sandwich": "sandwiches",
    "grab-go": "grab-go",
    "grab-and-go": "grab-go",
    "grab-go-hot-bar": "hot-bar",  # hot line — never also cold grab-go
    "grab-go-sandwiches": "sandwiches",
    "dips-and-spreads": "dips-and-spreads",
    "hot-bar": "hot-bar",
    "hb-menus": "hot-bar",
    "pizza": "pizza",
    "salads": "salads",
    "vegetarian": "vegetarian",
    "vegan": "vegan",
    "fall": "seasonal-fall",
    "winter": "seasonal-winter",
    "spring": "seasonal-spring",
    "summer": "seasonal-summer",
}

HB_DAY = {
    "hb1-sunday": "hot-bar-sun",
    "hb2-monday": "hot-bar-mon",
    "hb3-tuesday": "hot-bar-tue",
    "hb4-wednesday": "hot-bar-wed",
    "hb5-thursday": "hot-bar-thu",
    "hb6-friday": "hot-bar-fri",
    "hb7-saturday": "hot-bar-sat",
}

SKIP_NAME_RE = re.compile(
    r"^(<new recipe>|new recipe|test\b|xxx\b)", re.I
)
VARIANT_NOISE_RE = re.compile(
    r"\s*\((?:new recipe|new|old|test[^)]*|adjusted[^)]*)\)\s*$", re.I
)


def norm_unit(u: str | None) -> str | None:
    if u is None:
        return None
    key = u.strip().lower()
    if key in UNIT_MAP:
        return UNIT_MAP[key]
    # keep #10 can, bunch, head, jar as-is (kitchen units)
    return u.strip() or None


def clean_tags(tags: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for t in tags or []:
        t = (t or "").strip().lower()
        if not t or t == "work":
            continue
        if t in HB_DAY:
            mapped = HB_DAY[t]
        elif t in DEPT_ALIASES:
            mapped = DEPT_ALIASES[t]
        else:
            mapped = t
        if mapped not in seen:
            seen.add(mapped)
            out.append(mapped)
    # always keep work for deploy filter
    return ["work", *out]


def primary_dept(tags: list[str]) -> str:
    priority = [
        "bakery",
        "soups",
        "sandwiches",
        "pizza",
        "dips-and-spreads",
        "salads",
        "grab-and-go",
        "hot-bar",
        "vegan",
        "vegetarian",
    ]
    for p in priority:
        if p in tags:
            return p
    for t in tags:
        if t != "work" and not t.startswith("hot-bar-") and not t.startswith("seasonal-"):
            return t
    return "uncategorized"


def clean_recipe(r: dict) -> dict | None:
    name = (r.get("name") or "").strip()
    if not name or SKIP_NAME_RE.match(name):
        return None

    tags = clean_tags(r.get("tags") or [])
    ings_out = []
    for ing in r.get("ingredients") or []:
        unit = norm_unit(ing.get("unit"))
        ingredient = (ing.get("ingredient") or "").strip()
        if not ingredient:
            continue
        qty = ing.get("quantity")
        note = ing.get("note")
        # rebuild display
        parts = []
        if qty is not None and str(qty) != "":
            parts.append(str(qty))
        if unit:
            parts.append(unit)
        parts.append(ingredient)
        if note:
            parts.append(f"({note})")
        display = " ".join(parts)
        row = {
            "ingredient": ingredient,
            "display": display,
            "master_ingredient_name": ingredient,
        }
        if qty is not None and str(qty) != "":
            row["quantity"] = str(qty)
        if unit:
            row["unit"] = unit
        if note:
            row["note"] = note
        ings_out.append(row)

    steps = []
    for s in r.get("steps") or []:
        instr = (s.get("instruction") or "").strip()
        if not instr or instr.lower().rstrip(":") in {
            "instructions",
            "assembly",
            "procedure",
            "method",
            "directions",
        }:
            continue
        steps.append(
            {
                "position": len(steps) + 1,
                "instruction": instr,
            }
        )

    out = {
        "name": name,
        "description": r.get("description"),
        "servings": r.get("servings") or 1,
        "prep_time_minutes": r.get("prep_time_minutes"),
        "cook_time_minutes": r.get("cook_time_minutes"),
        "total_time_minutes": r.get("total_time_minutes"),
        "author": r.get("author"),
        "yield_quantity": r.get("yield_quantity"),
        "yield_unit": norm_unit(r.get("yield_unit")) if r.get("yield_unit") else r.get("yield_unit"),
        "tags": tags,
        "ingredients": ings_out,
        "steps": steps,
    }
    # drop empties
    return {k: v for k, v in out.items() if v is not None and v != []}


def recipe_to_md(r: dict) -> str:
    lines = [f"# {r['name']}", ""]
    meta = []
    if r.get("servings"):
        meta.append(f"**Servings:** {r['servings']}")
    if r.get("yield_quantity"):
        meta.append(
            f"**Yield:** {r['yield_quantity']} {r.get('yield_unit') or ''}".strip()
        )
    if r.get("prep_time_minutes"):
        meta.append(f"**Prep:** {r['prep_time_minutes']} min")
    if r.get("cook_time_minutes"):
        meta.append(f"**Cook:** {r['cook_time_minutes']} min")
    if r.get("author"):
        meta.append(f"**Author:** {r['author']}")
    if r.get("tags"):
        meta.append("**Tags:** " + ", ".join(f"`{t}`" for t in r["tags"] if t != "work"))
    if meta:
        lines.append(" · ".join(meta))
        lines.append("")
    if r.get("description"):
        lines.append(f"_{r['description']}_")
        lines.append("")
    lines.append("## Ingredients")
    for ing in r.get("ingredients") or []:
        lines.append(f"- {ing.get('display') or ing.get('ingredient')}")
    lines.append("")
    lines.append("## Method")
    for s in r.get("steps") or []:
        lines.append(f"{s['position']}. {s['instruction']}")
    if not r.get("steps"):
        lines.append("_No procedure in source export._")
    lines.append("")
    lines.append("---")
    lines.append("")
    return "\n".join(lines)


def export_sandwich_board(out_dir: Path) -> None:
    xlsx = SRC / "UPDATED Sandwich Daily Prep 2026.xlsx"
    if not xlsx.is_file():
        return
    try:
        from openpyxl import load_workbook
    except ImportError:
        return
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    path = out_dir / "sandwich-board.csv"
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(
            [
                "sandwich",
                "am_par",
                "pm_par",
                "ovnt_par",
                "prep_task",
                "prep_par",
            ]
        )
        for row in rows[2:]:
            if not row:
                continue
            sandwich = row[1]
            prep = row[12] if len(row) > 12 else None
            if not sandwich and not prep:
                continue
            w.writerow(
                [
                    sandwich or "",
                    row[2] if len(row) > 2 else "",
                    row[5] if len(row) > 5 else "",
                    row[8] if len(row) > 8 else "",
                    prep or "",
                    row[14] if len(row) > 14 else "",
                ]
            )
    print(f"  sandwich-board.csv", file=sys.stderr)


def export_catalog(out_dir: Path, imported_names: set[str]) -> None:
    src = SRC / "RecpList.txt"
    if not src.is_file():
        return
    path = out_dir / "catalog.csv"
    with src.open(encoding="latin-1", newline="") as fin, path.open(
        "w", newline="", encoding="utf-8"
    ) as fout:
        reader = csv.DictReader(fin)
        fields = list(reader.fieldnames or []) + ["in_larder"]
        w = csv.DictWriter(fout, fieldnames=fields)
        w.writeheader()
        for row in reader:
            name = (row.get("Recipe") or "").strip()
            row["in_larder"] = "yes" if name.casefold() in imported_names else "no"
            w.writerow(row)
    print(f"  catalog.csv", file=sys.stderr)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--bundle",
        type=Path,
        default=SRC / "lexco-bundle.json",
        help="Source full bundle JSON",
    )
    ap.add_argument(
        "--out",
        type=Path,
        default=OUT,
        help="Output directory",
    )
    args = ap.parse_args()

    if not args.bundle.is_file():
        print(f"error: missing {args.bundle}", file=sys.stderr)
        return 1

    print(f"Loading {args.bundle} …", file=sys.stderr)
    bundle = json.loads(args.bundle.read_text(encoding="utf-8"))
    raw_recipes = bundle.get("recipes") or []

    cleaned: list[dict] = []
    skipped = 0
    for r in raw_recipes:
        c = clean_recipe(r)
        if c is None:
            skipped += 1
            continue
        cleaned.append(c)

    print(f"  cleaned recipes: {len(cleaned)} (skipped {skipped})", file=sys.stderr)

    out = args.out
    out.mkdir(parents=True, exist_ok=True)
    by_dept = out / "by-dept"
    by_dept.mkdir(exist_ok=True)

    # masters
    master_counts: Counter[str] = Counter()
    master_unit: dict[str, str | None] = {}
    for r in cleaned:
        for ing in r.get("ingredients") or []:
            name = ing["ingredient"]
            key = name.casefold()
            master_counts[key] += 1
            if key not in master_unit and ing.get("unit"):
                master_unit[key] = ing["unit"]
            elif key not in master_unit:
                master_unit[key] = None

    masters = []
    name_by_key = {}
    for r in cleaned:
        for ing in r.get("ingredients") or []:
            name_by_key.setdefault(ing["ingredient"].casefold(), ing["ingredient"])
    for key, count in master_counts.most_common():
        m = {"name": name_by_key[key], "usage_count": count}
        if master_unit.get(key):
            m["default_unit"] = master_unit[key]
        masters.append(m)

    kitchen_bundle = {
        "larder_version": 1,
        "exported_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "ingredients": [
            {
                "name": m["name"],
                **({"default_unit": m["default_unit"]} if m.get("default_unit") else {}),
            }
            for m in masters
        ],
        "location_prices": [],
        "recipes": cleaned,
    }
    (out / "kitchen-bundle.json").write_text(
        json.dumps(kitchen_bundle, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    print(
        f"  kitchen-bundle.json ({(out / 'kitchen-bundle.json').stat().st_size // 1024} KB)",
        file=sys.stderr,
    )

    # recipes.csv
    with (out / "recipes.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(
            [
                "recipe",
                "dept",
                "servings",
                "prep_min",
                "cook_min",
                "yield_qty",
                "yield_unit",
                "author",
                "tags",
                "line_type",
                "quantity",
                "unit",
                "ingredient",
                "note",
                "display_or_step",
            ]
        )
        for r in cleaned:
            dept = primary_dept(r.get("tags") or [])
            tags = "|".join(t for t in r.get("tags") or [] if t != "work")
            base = [
                r["name"],
                dept,
                r.get("servings") or "",
                r.get("prep_time_minutes") or "",
                r.get("cook_time_minutes") or "",
                r.get("yield_quantity") or "",
                r.get("yield_unit") or "",
                r.get("author") or "",
                tags,
            ]
            if not r.get("ingredients") and not r.get("steps"):
                w.writerow(base + ["empty", "", "", "", "", ""])
            for ing in r.get("ingredients") or []:
                w.writerow(
                    base
                    + [
                        "ingredient",
                        ing.get("quantity") or "",
                        ing.get("unit") or "",
                        ing.get("ingredient") or "",
                        ing.get("note") or "",
                        ing.get("display") or "",
                    ]
                )
            for s in r.get("steps") or []:
                w.writerow(
                    base
                    + [
                        "step",
                        s.get("position") or "",
                        "",
                        "",
                        "",
                        s.get("instruction") or "",
                    ]
                )
    print("  recipes.csv", file=sys.stderr)

    with (out / "ingredients-master.csv").open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f, fieldnames=["name", "default_unit", "usage_count"]
        )
        w.writeheader()
        for m in masters:
            w.writerow(
                {
                    "name": m["name"],
                    "default_unit": m.get("default_unit") or "",
                    "usage_count": m["usage_count"],
                }
            )
    print(f"  ingredients-master.csv ({len(masters)} rows)", file=sys.stderr)

    # by department
    groups: dict[str, list[dict]] = defaultdict(list)
    for r in cleaned:
        groups[primary_dept(r.get("tags") or [])].append(r)

    index_lines = [
        "# Lexington Co-op — cleaned kitchen exports",
        "",
        f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
        f"Recipes: **{len(cleaned)}** · Ingredients: **{len(masters)}**",
        "",
        "## Files",
        "",
        "| File | Purpose |",
        "| --- | --- |",
        "| `kitchen-bundle.json` | Re-import into larder (`import --file`) |",
        "| `recipes.csv` | Spreadsheet: one row per ingredient/step |",
        "| `ingredients-master.csv` | Unique ingredients + usage |",
        "| `catalog.csv` | Full ChefTec list vs imported |",
        "| `sandwich-board.csv` | Daily sandwich pars / prep |",
        "| `by-dept/*.md` | Printable department cookbooks |",
        "",
        "## Departments",
        "",
    ]

    for dept in sorted(groups.keys(), key=lambda d: (-len(groups[d]), d)):
        recs = sorted(groups[dept], key=lambda r: r["name"].casefold())
        md_path = by_dept / f"{dept}.md"
        json_path = by_dept / f"{dept}.json"
        body = [f"# {dept.replace('-', ' ').title()}", "", f"_{len(recs)} recipes_", ""]
        for r in recs:
            body.append(recipe_to_md(r))
        md_path.write_text("\n".join(body), encoding="utf-8")
        json_path.write_text(
            json.dumps(recs, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
        )
        index_lines.append(f"- **{dept}** — {len(recs)} → [`by-dept/{dept}.md`](by-dept/{dept}.md)")
        print(f"  by-dept/{dept}.md ({len(recs)})", file=sys.stderr)

    (out / "index.md").write_text("\n".join(index_lines) + "\n", encoding="utf-8")

    imported = {r["name"].casefold() for r in cleaned}
    export_catalog(out, imported)
    export_sandwich_board(out)

    print(f"Done → {out}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
