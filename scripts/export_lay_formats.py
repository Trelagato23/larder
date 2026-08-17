#!/usr/bin/env python3
"""Build lay-person friendly kitchen packs from cleaned kitchen-bundle.json.

Outputs under data/lexco-source/clean/for-everyone/:
  START-HERE.html          open this first (no install)
  recipe-browser.html      searchable offline recipe book
  recipes-data.js          data for the browser
  kitchen.xlsx             Excel workbook (sheets by dept + lists)
  print/<dept>.html        print-friendly pages (File → Print → PDF)
  README.txt               plain-English guide

Usage:
  python3 scripts/export_lay_formats.py
"""

from __future__ import annotations

import argparse
import html
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None  # type: ignore

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_BUNDLE = ROOT / "data" / "lexco-source" / "clean" / "kitchen-bundle.json"
DEFAULT_OUT = ROOT / "data" / "lexco-source" / "clean" / "for-everyone"

DEPT_PRIORITY = [
    "bakery",
    "soups",
    "sandwiches",
    "pizza",
    "dips-and-spreads",
    "salads",
    "grab-and-go",
    "hot-bar",
    "dressing",
    "vegan",
    "vegetarian",
    "uncategorized",
]


def primary_dept(tags: list[str]) -> str:
    tags = [t for t in (tags or []) if t != "work"]
    for p in DEPT_PRIORITY:
        if p in tags:
            return p
    return tags[0] if tags else "uncategorized"


def title_dept(d: str) -> str:
    return d.replace("-", " ").title()


def slug(s: str) -> str:
    s = s.casefold()
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s[:80] or "recipe"


def load_recipes(bundle_path: Path) -> list[dict]:
    data = json.loads(bundle_path.read_text(encoding="utf-8"))
    recipes = data.get("recipes") or []
    for r in recipes:
        r["_dept"] = primary_dept(r.get("tags") or [])
    recipes.sort(key=lambda r: (r["_dept"], (r.get("name") or "").casefold()))
    return recipes


def write_excel(recipes: list[dict], path: Path) -> None:
    if Workbook is None:
        print("  skip kitchen.xlsx (openpyxl missing)", file=sys.stderr)
        return
    wb = Workbook()

    # --- All recipes sheet (one row per recipe summary) ---
    ws = wb.active
    ws.title = "All recipes"
    headers = [
        "Recipe",
        "Department",
        "Servings",
        "Prep (min)",
        "Cook (min)",
        "Yield",
        "Author",
        "Tags",
        "Ingredient count",
        "Has method?",
    ]
    header_fill = PatternFill("solid", fgColor="B42318")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
    for i, r in enumerate(recipes, 2):
        yq = r.get("yield_quantity") or ""
        yu = r.get("yield_unit") or ""
        yield_s = f"{yq} {yu}".strip()
        tags = ", ".join(t for t in (r.get("tags") or []) if t != "work")
        ws.cell(i, 1, r.get("name"))
        ws.cell(i, 2, title_dept(r["_dept"]))
        ws.cell(i, 3, r.get("servings") or "")
        ws.cell(i, 4, r.get("prep_time_minutes") or "")
        ws.cell(i, 5, r.get("cook_time_minutes") or "")
        ws.cell(i, 6, yield_s)
        ws.cell(i, 7, r.get("author") or "")
        ws.cell(i, 8, tags)
        ws.cell(i, 9, len(r.get("ingredients") or []))
        ws.cell(i, 10, "yes" if r.get("steps") else "no")
    ws.auto_filter.ref = f"A1:J{len(recipes) + 1}"
    ws.freeze_panes = "A2"
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["H"].width = 28

    # --- Ingredients lines sheet ---
    ws2 = wb.create_sheet("Ingredient lines")
    h2 = ["Recipe", "Department", "Qty", "Unit", "Ingredient", "Note", "Full line"]
    for col, h in enumerate(h2, 1):
        cell = ws2.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
    row = 2
    for r in recipes:
        for ing in r.get("ingredients") or []:
            ws2.cell(row, 1, r.get("name"))
            ws2.cell(row, 2, title_dept(r["_dept"]))
            ws2.cell(row, 3, ing.get("quantity") or "")
            ws2.cell(row, 4, ing.get("unit") or "")
            ws2.cell(row, 5, ing.get("ingredient") or "")
            ws2.cell(row, 6, ing.get("note") or "")
            ws2.cell(row, 7, ing.get("display") or "")
            row += 1
    ws2.auto_filter.ref = f"A1:G{max(row - 1, 1)}"
    ws2.freeze_panes = "A2"
    for col, w in enumerate([36, 16, 10, 10, 28, 22, 40], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w

    # --- Method steps sheet ---
    ws3 = wb.create_sheet("Method steps")
    h3 = ["Recipe", "Department", "Step #", "Instruction"]
    for col, h in enumerate(h3, 1):
        cell = ws3.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
    row = 2
    for r in recipes:
        for s in r.get("steps") or []:
            ws3.cell(row, 1, r.get("name"))
            ws3.cell(row, 2, title_dept(r["_dept"]))
            ws3.cell(row, 3, s.get("position") or "")
            cell = ws3.cell(row, 4, s.get("instruction") or "")
            cell.alignment = Alignment(wrap_text=True)
            row += 1
    ws3.auto_filter.ref = f"A1:D{max(row - 1, 1)}"
    ws3.freeze_panes = "A2"
    ws3.column_dimensions["A"].width = 36
    ws3.column_dimensions["B"].width = 16
    ws3.column_dimensions["C"].width = 8
    ws3.column_dimensions["D"].width = 80

    # --- One sheet per major dept (recipe cards as blocks via rows) ---
    by_dept: dict[str, list] = defaultdict(list)
    for r in recipes:
        by_dept[r["_dept"]].append(r)

    for dept in sorted(by_dept.keys(), key=lambda d: (-len(by_dept[d]), d)):
        name = title_dept(dept)[:31]
        wsd = wb.create_sheet(name)
        wsd.cell(1, 1, f"{title_dept(dept)} — {len(by_dept[dept])} recipes")
        wsd["A1"].font = Font(bold=True, size=14, color="B42318")
        wsd.merge_cells("A1:D1")
        rnum = 3
        for r in by_dept[dept]:
            wsd.cell(rnum, 1, r.get("name")).font = Font(bold=True, size=12)
            rnum += 1
            meta = []
            if r.get("servings"):
                meta.append(f"Servings: {r['servings']}")
            if r.get("prep_time_minutes"):
                meta.append(f"Prep: {r['prep_time_minutes']} min")
            if r.get("cook_time_minutes"):
                meta.append(f"Cook: {r['cook_time_minutes']} min")
            if r.get("yield_quantity"):
                meta.append(
                    f"Yield: {r.get('yield_quantity')} {r.get('yield_unit') or ''}".strip()
                )
            if meta:
                wsd.cell(rnum, 1, " · ".join(meta))
                rnum += 1
            wsd.cell(rnum, 1, "Ingredients:").font = Font(bold=True)
            rnum += 1
            for ing in r.get("ingredients") or []:
                wsd.cell(rnum, 1, f"  • {ing.get('display') or ing.get('ingredient')}")
                rnum += 1
            wsd.cell(rnum, 1, "Method:").font = Font(bold=True)
            rnum += 1
            for s in r.get("steps") or []:
                cell = wsd.cell(
                    rnum,
                    1,
                    f"  {s.get('position')}. {s.get('instruction')}",
                )
                cell.alignment = Alignment(wrap_text=True)
                rnum += 1
            if not r.get("steps"):
                wsd.cell(rnum, 1, "  (No written method in source)")
                rnum += 1
            rnum += 1  # blank between recipes
        wsd.column_dimensions["A"].width = 100

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"  kitchen.xlsx ({path.stat().st_size // 1024} KB)", file=sys.stderr)


def compact_for_js(recipes: list[dict]) -> list[dict]:
    out = []
    for r in recipes:
        out.append(
            {
                "id": slug(r.get("name") or "x"),
                "name": r.get("name"),
                "dept": r["_dept"],
                "deptLabel": title_dept(r["_dept"]),
                "servings": r.get("servings"),
                "prep": r.get("prep_time_minutes"),
                "cook": r.get("cook_time_minutes"),
                "yield": " ".join(
                    str(x)
                    for x in (r.get("yield_quantity"), r.get("yield_unit"))
                    if x
                ).strip()
                or None,
                "author": r.get("author"),
                "desc": r.get("description"),
                "tags": [t for t in (r.get("tags") or []) if t != "work"],
                "ings": [
                    i.get("display") or i.get("ingredient")
                    for i in (r.get("ingredients") or [])
                ],
                "steps": [
                    s.get("instruction")
                    for s in (r.get("steps") or [])
                    if s.get("instruction")
                ],
            }
        )
    return out


BROWSER_HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>Lexington Co-op Kitchen Recipes</title>
<style>
  :root {
    --brand: #b42318;
    --ink: #1c1917;
    --muted: #57534e;
    --paper: #faf7f2;
    --card: #fffdf8;
    --line: #e7e0d5;
    --accent: #0f766e;
  }
  * { box-sizing: border-box; }
  body {
    margin: 0;
    font-family: "Segoe UI", system-ui, -apple-system, sans-serif;
    background: var(--paper);
    color: var(--ink);
    line-height: 1.45;
  }
  header {
    background: var(--brand);
    color: #fff;
    padding: 1rem 1.25rem;
    position: sticky; top: 0; z-index: 10;
    box-shadow: 0 2px 10px rgba(0,0,0,.15);
  }
  header h1 { margin: 0; font-size: 1.25rem; font-weight: 700; }
  header p { margin: .25rem 0 0; opacity: .9; font-size: .9rem; }
  .wrap { max-width: 1100px; margin: 0 auto; padding: 1rem; }
  .controls {
    display: grid;
    gap: .75rem;
    grid-template-columns: 1fr;
    margin-bottom: 1rem;
  }
  @media (min-width: 720px) {
    .controls { grid-template-columns: 2fr 1fr auto; align-items: end; }
  }
  label { display: block; font-size: .8rem; color: var(--muted); margin-bottom: .25rem; }
  input[type=search], select {
    width: 100%;
    padding: .65rem .75rem;
    border: 1px solid var(--line);
    border-radius: 8px;
    font-size: 1rem;
    background: #fff;
  }
  button {
    padding: .65rem 1rem;
    border: 0;
    border-radius: 8px;
    background: var(--brand);
    color: #fff;
    font-weight: 600;
    cursor: pointer;
  }
  button.secondary { background: #44403c; }
  .meta { color: var(--muted); font-size: .9rem; margin: .5rem 0 1rem; }
  .list { display: grid; gap: .5rem; }
  .card {
    background: var(--card);
    border: 1px solid var(--line);
    border-left: 5px solid var(--brand);
    border-radius: 10px;
    padding: .85rem 1rem;
    cursor: pointer;
  }
  .card:hover { border-color: var(--brand); }
  .card h2 { margin: 0 0 .25rem; font-size: 1.05rem; }
  .card .sub { color: var(--muted); font-size: .85rem; }
  .badge {
    display: inline-block;
    background: #f5e6e4;
    color: var(--brand);
    font-size: .75rem;
    font-weight: 600;
    padding: .15rem .45rem;
    border-radius: 999px;
    margin-right: .25rem;
  }
  /* detail */
  #detail { display: none; }
  #detail.open { display: block; }
  #list-view.hide { display: none; }
  .back { margin-bottom: 1rem; }
  .recipe-sheet {
    background: var(--card);
    border: 1px solid var(--line);
    border-radius: 12px;
    padding: 1.25rem 1.5rem;
  }
  .recipe-sheet h1 { margin-top: 0; color: var(--brand); }
  .facts { display: flex; flex-wrap: wrap; gap: .5rem 1rem; color: var(--muted); margin-bottom: 1rem; }
  .recipe-sheet h3 { margin: 1.25rem 0 .5rem; border-bottom: 1px solid var(--line); padding-bottom: .25rem; }
  ul.ings { padding-left: 1.2rem; }
  ol.steps { padding-left: 1.2rem; }
  ul.ings li, ol.steps li { margin: .35rem 0; }
  .print-only { display: none; }
  @media print {
    header, .controls, .back, .meta, #list-view { display: none !important; }
    #detail, #detail.open { display: block !important; }
    .print-only { display: block; }
    body { background: #fff; }
    .recipe-sheet { border: 0; padding: 0; }
  }
</style>
</head>
<body>
<header>
  <h1>Lexington Co-op Kitchen Recipes</h1>
  <p>Offline recipe book — search, filter, print. No internet needed.</p>
</header>
<div class="wrap">
  <div id="list-view">
    <div class="controls">
      <div>
        <label for="q">Search recipes or ingredients</label>
        <input id="q" type="search" placeholder="e.g. hummus, tofu, bakery…" autocomplete="off"/>
      </div>
      <div>
        <label for="dept">Department</label>
        <select id="dept"><option value="">All departments</option></select>
      </div>
      <div>
        <label>&nbsp;</label>
        <button type="button" class="secondary" id="clear">Clear</button>
      </div>
    </div>
    <p class="meta" id="count"></p>
    <div class="list" id="list"></div>
  </div>

  <div id="detail">
    <button type="button" class="back" id="back">← Back to list</button>
    <button type="button" class="back" id="printBtn" style="margin-left:.5rem">Print this recipe</button>
    <div class="recipe-sheet" id="sheet"></div>
  </div>
</div>
<script src="recipes-data.js"></script>
<script>
(function () {
  const data = window.LEXCO_RECIPES || [];
  const listEl = document.getElementById('list');
  const countEl = document.getElementById('count');
  const qEl = document.getElementById('q');
  const deptEl = document.getElementById('dept');
  const listView = document.getElementById('list-view');
  const detail = document.getElementById('detail');
  const sheet = document.getElementById('sheet');
  let active = null;

  const depts = [...new Set(data.map(r => r.dept))].sort();
  for (const d of depts) {
    const opt = document.createElement('option');
    opt.value = d;
    opt.textContent = data.find(r => r.dept === d).deptLabel + ' (' + data.filter(r => r.dept === d).length + ')';
    deptEl.appendChild(opt);
  }

  function esc(s) {
    return String(s == null ? '' : s)
      .replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
  }

  function matches(r, q, dept) {
    if (dept && r.dept !== dept) return false;
    if (!q) return true;
    const hay = [
      r.name, r.deptLabel, r.author, r.desc,
      ...(r.tags || []), ...(r.ings || []), ...(r.steps || [])
    ].join(' ').toLowerCase();
    return hay.includes(q);
  }

  function renderList() {
    const q = (qEl.value || '').trim().toLowerCase();
    const dept = deptEl.value;
    const rows = data.filter(r => matches(r, q, dept));
    countEl.textContent = rows.length + ' recipe' + (rows.length === 1 ? '' : 's')
      + (q || dept ? ' matching' : ' total');
    listEl.innerHTML = rows.slice(0, 400).map(r => {
      const bits = [];
      if (r.servings) bits.push(r.servings + ' servings');
      if (r.prep) bits.push(r.prep + ' min prep');
      if (r.cook) bits.push(r.cook + ' min cook');
      return '<article class="card" data-id="' + esc(r.id) + '">'
        + '<span class="badge">' + esc(r.deptLabel) + '</span>'
        + '<h2>' + esc(r.name) + '</h2>'
        + '<div class="sub">' + esc(bits.join(' · ')) + '</div>'
        + '</article>';
    }).join('');
    if (rows.length > 400) {
      listEl.innerHTML += '<p class="meta">Showing first 400 — refine your search.</p>';
    }
  }

  function showRecipe(r) {
    active = r;
    listView.classList.add('hide');
    detail.classList.add('open');
    const facts = [];
    if (r.servings) facts.push('<strong>Servings</strong> ' + esc(r.servings));
    if (r.yield) facts.push('<strong>Yield</strong> ' + esc(r.yield));
    if (r.prep) facts.push('<strong>Prep</strong> ' + esc(r.prep) + ' min');
    if (r.cook) facts.push('<strong>Cook</strong> ' + esc(r.cook) + ' min');
    if (r.author) facts.push('<strong>Author</strong> ' + esc(r.author));
    const ings = (r.ings || []).map(i => '<li>' + esc(i) + '</li>').join('') || '<li><em>No ingredients listed</em></li>';
    const steps = (r.steps || []).map(s => '<li>' + esc(s) + '</li>').join('') || '<li><em>No method written in source file</em></li>';
    sheet.innerHTML =
      '<p class="print-only"><strong>Lexington Co-op Kitchen</strong></p>'
      + '<h1>' + esc(r.name) + '</h1>'
      + '<p><span class="badge">' + esc(r.deptLabel) + '</span></p>'
      + (r.desc ? '<p class="sub">' + esc(r.desc) + '</p>' : '')
      + '<div class="facts">' + facts.map(f => '<span>' + f + '</span>').join('') + '</div>'
      + '<h3>Ingredients</h3><ul class="ings">' + ings + '</ul>'
      + '<h3>Method</h3><ol class="steps">' + steps + '</ol>';
    window.scrollTo(0, 0);
  }

  listEl.addEventListener('click', (e) => {
    const card = e.target.closest('.card');
    if (!card) return;
    const r = data.find(x => x.id === card.dataset.id);
    if (r) showRecipe(r);
  });
  document.getElementById('back').onclick = () => {
    detail.classList.remove('open');
    listView.classList.remove('hide');
    active = null;
  };
  document.getElementById('printBtn').onclick = () => window.print();
  document.getElementById('clear').onclick = () => {
    qEl.value = ''; deptEl.value = ''; renderList();
  };
  qEl.addEventListener('input', renderList);
  deptEl.addEventListener('change', renderList);
  renderList();
})();
</script>
</body>
</html>
"""

START_HERE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>Lex Kitchen Recipes — Start Here</title>
<style>
  body { font-family: "Segoe UI", system-ui, sans-serif; max-width: 720px; margin: 2rem auto; padding: 0 1rem; line-height: 1.5; color: #1c1917; background: #faf7f2; }
  h1 { color: #b42318; }
  a.card { display: block; background: #fff; border: 1px solid #e7e0d5; border-left: 6px solid #b42318; border-radius: 10px; padding: 1rem 1.2rem; margin: .75rem 0; text-decoration: none; color: inherit; }
  a.card:hover { border-color: #b42318; }
  a.card strong { color: #b42318; font-size: 1.1rem; }
  .muted { color: #57534e; }
  code { background: #f5f5f4; padding: .1rem .35rem; border-radius: 4px; }
</style>
</head>
<body>
  <h1>Lexington Co-op kitchen recipes</h1>
  <p class="muted">Cleaned recipe pack for kitchen &amp; office use. No install required for the options below.</p>

  <a class="card" href="recipe-browser.html">
    <strong>1. Recipe browser (easiest)</strong><br/>
    Double-click this. Search by name or ingredient, filter by department, print one recipe.
  </a>

  <a class="card" href="kitchen.xlsx">
    <strong>2. Excel workbook</strong><br/>
    Open in Excel / Google Sheets / LibreOffice. Filterable lists + full sheets per department.
  </a>

  <a class="card" href="print/">
    <strong>3. Print packets by department</strong><br/>
    Open a department page → File → Print → “Save as PDF” for a paper booklet.
  </a>

  <h2>Tips for staff</h2>
  <ul>
    <li>Prefer the <strong>recipe browser</strong> on a kitchen tablet or laptop.</li>
    <li>Use <strong>Excel</strong> for inventory / costing / lists (office).</li>
    <li>Use <strong>Print → PDF</strong> when someone needs a paper prep sheet.</li>
    <li>These files are a snapshot of the kitchen database — ask a manager if something looks outdated.</li>
  </ul>

  <p class="muted">Also on this USB (parent folders): CSV files, markdown cookbooks, and technical JSON for IT.</p>
</body>
</html>
"""


def write_print_dept(dept: str, recipes: list[dict], path: Path) -> None:
    parts = [
        "<!DOCTYPE html><html lang='en'><head><meta charset='utf-8'/>",
        f"<title>{html.escape(title_dept(dept))} — Lex Kitchen</title>",
        """<style>
        body{font-family:Georgia,serif;max-width:800px;margin:1rem auto;padding:0 1rem;color:#111}
        h1{color:#b42318;border-bottom:3px solid #b42318;padding-bottom:.3rem}
        h2{page-break-before:always;color:#b42318;margin-top:2rem}
        h2:first-of-type{page-break-before:avoid}
        .meta{color:#444;font-size:.95rem;margin:.25rem 0 1rem}
        ul{margin:.3rem 0 1rem}
        ol{margin:.3rem 0}
        li{margin:.25rem 0}
        .toc a{color:#0f766e}
        @media print{a{color:inherit;text-decoration:none}}
        </style></head><body>""",
        f"<h1>{html.escape(title_dept(dept))}</h1>",
        f"<p class='meta'>{len(recipes)} recipes · Lexington Co-op · print via browser (Ctrl+P)</p>",
        "<div class='toc'><h3>Contents</h3><ol>",
    ]
    for r in recipes:
        rid = slug(r.get("name") or "x")
        parts.append(
            f"<li><a href='#{html.escape(rid)}'>{html.escape(r.get('name') or '')}</a></li>"
        )
    parts.append("</ol></div>")

    for r in recipes:
        rid = slug(r.get("name") or "x")
        parts.append(f"<h2 id='{html.escape(rid)}'>{html.escape(r.get('name') or '')}</h2>")
        meta = []
        if r.get("servings"):
            meta.append(f"Servings: {r['servings']}")
        if r.get("prep_time_minutes"):
            meta.append(f"Prep: {r['prep_time_minutes']} min")
        if r.get("cook_time_minutes"):
            meta.append(f"Cook: {r['cook_time_minutes']} min")
        if r.get("yield_quantity"):
            meta.append(
                f"Yield: {r.get('yield_quantity')} {r.get('yield_unit') or ''}".strip()
            )
        if r.get("author"):
            meta.append(f"Author: {r['author']}")
        if meta:
            parts.append(f"<p class='meta'>{html.escape(' · '.join(meta))}</p>")
        if r.get("description"):
            parts.append(f"<p><em>{html.escape(r['description'])}</em></p>")
        parts.append("<h3>Ingredients</h3><ul>")
        for ing in r.get("ingredients") or []:
            line = ing.get("display") or ing.get("ingredient") or ""
            parts.append(f"<li>{html.escape(line)}</li>")
        if not r.get("ingredients"):
            parts.append("<li><em>None listed</em></li>")
        parts.append("</ul><h3>Method</h3><ol>")
        for s in r.get("steps") or []:
            parts.append(f"<li>{html.escape(s.get('instruction') or '')}</li>")
        if not r.get("steps"):
            parts.append("<li><em>No method in source export</em></li>")
        parts.append("</ol>")

    parts.append("</body></html>")
    path.write_text("\n".join(parts), encoding="utf-8")


def write_print_index(depts: dict[str, list], path: Path) -> None:
    lines = [
        "<!DOCTYPE html><html><head><meta charset='utf-8'/><title>Print packets</title>",
        "<style>body{font-family:system-ui,sans-serif;max-width:640px;margin:2rem auto;padding:0 1rem}",
        "a{display:block;padding:.6rem .8rem;margin:.4rem 0;background:#fff;border:1px solid #ddd;",
        "border-left:5px solid #b42318;text-decoration:none;color:#111;border-radius:8px}</style></head><body>",
        "<h1>Print packets by department</h1>",
        "<p>Open a link, then use <strong>File → Print → Save as PDF</strong> (or print paper).</p>",
    ]
    for dept in sorted(depts.keys(), key=lambda d: (-len(depts[d]), d)):
        lines.append(
            f"<a href='{html.escape(dept)}.html'><strong>{html.escape(title_dept(dept))}</strong>"
            f" — {len(depts[dept])} recipes</a>"
        )
    lines.append("</body></html>")
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--bundle", type=Path, default=DEFAULT_BUNDLE)
    ap.add_argument("--out", type=Path, default=DEFAULT_OUT)
    args = ap.parse_args()
    if not args.bundle.is_file():
        print(f"error: missing {args.bundle}", file=sys.stderr)
        return 1

    print(f"Loading {args.bundle} …", file=sys.stderr)
    recipes = load_recipes(args.bundle)
    print(f"  {len(recipes)} recipes", file=sys.stderr)
    out = args.out
    out.mkdir(parents=True, exist_ok=True)
    print_dir = out / "print"
    print_dir.mkdir(exist_ok=True)

    # Excel
    write_excel(recipes, out / "kitchen.xlsx")

    # Browser data + HTML
    compact = compact_for_js(recipes)
    (out / "recipes-data.js").write_text(
        "window.LEXCO_RECIPES = "
        + json.dumps(compact, ensure_ascii=False, separators=(",", ":"))
        + ";\n",
        encoding="utf-8",
    )
    print(
        f"  recipes-data.js ({(out / 'recipes-data.js').stat().st_size // 1024} KB)",
        file=sys.stderr,
    )
    (out / "recipe-browser.html").write_text(BROWSER_HTML, encoding="utf-8")
    (out / "START-HERE.html").write_text(START_HERE, encoding="utf-8")
    print("  START-HERE.html, recipe-browser.html", file=sys.stderr)

    # Print packets
    by_dept: dict[str, list] = defaultdict(list)
    for r in recipes:
        by_dept[r["_dept"]].append(r)
    for dept, recs in by_dept.items():
        write_print_dept(dept, recs, print_dir / f"{dept}.html")
    write_print_index(by_dept, print_dir / "index.html")
    print(f"  print/ ({len(by_dept)} department packets)", file=sys.stderr)

    readme = out / "README.txt"
    readme.write_text(
        """Lexington Co-op recipes — for kitchen & office staff
=====================================================

START HERE
  Double-click:  START-HERE.html

1) RECIPE BROWSER (tablets / laptops)
   Open recipe-browser.html
   - Type to search (name or ingredient)
   - Filter by department
   - Click a recipe → Print for a paper copy

2) EXCEL
   Open kitchen.xlsx
   - "All recipes" = searchable list with filters
   - "Ingredient lines" / "Method steps" = full detail
   - Other tabs = one department each (readable cards)

3) PRINT / PDF BOOKLETS
   Open print/index.html → pick a department
   Browser menu: File → Print → Save as PDF
   (Works on Windows, Mac, Chromebooks — no special software)

No internet required. Keep this folder together (don't move
recipe-browser.html without recipes-data.js).
""",
        encoding="utf-8",
    )
    print(f"Done → {out}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
